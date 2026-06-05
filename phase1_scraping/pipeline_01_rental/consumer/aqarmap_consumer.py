"""
Aqarmap Rental Data — Kafka Consumer + Cleaner
================================================
يستقبل events من الـ Producer
→ ينظف الداتا
→ يحفظها في Excel بنفس شكل rent_commercial3_cleaned.xlsx

Output columns:
    area(m²)     | int
    rent(EGP)    | int
    location_en  | str  →  "Alexandria / Neighborhood"
"""

import json
import logging
import os
import re
from datetime import datetime, timezone

import pandas as pd
from kafka import KafkaConsumer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [CONSUMER] %(message)s")
log = logging.getLogger(__name__)

# ─── Config ────────────────────────────────────────────────────────────────────
KAFKA_BROKER   = os.environ.get("KAFKA_BOOTSTRAP_SERVERS", "kafka:9092")
TOPIC          = "rent-commercial-alexandria"
GROUP_ID       = "rent-cleaner-group"
OUTPUT_DIR     = os.path.join(os.path.dirname(__file__), "..", "output")
BATCH_TIMEOUT  = 60_000      # ms — flush batch لو مفيش messages جديدة
# ───────────────────────────────────────────────────────────────────────────────


# ─── Cleaning Rules (تطابق clean_data_rent.ipynb) ──────────────────────────────

AREA_MIN, AREA_MAX   = 50, 5000      # م² — نفس range الداتا الأصلية
RENT_MIN, RENT_MAX   = 1100, 700_000  # ج.م — نفس range الداتا الأصلية

LOCATION_FIXES = {               # تصحيح أخطاء إملائية شائعة من الـ scraper
    "Moharam Bek":   "Moharram Bek",
    "Moharam Bey":   "Moharram Bek",
    "El Asafra":     "Asafra Bahary",
    "Kafr Abdu":     "Kafr Abdo",
    "San Stefanus":  "San Stefano",
}


def clean_location(raw: str) -> str | None:
    """
    'Alexandria / smouha'  →  'Alexandria / Smouha'
    يتأكد من البداية بـ 'Alexandria'، ويعمل Title Case
    """
    if not raw or "/" not in raw:
        return None
    parts = [p.strip().title() for p in raw.split("/")]
    if parts[0].lower() not in ("alexandria", "al iskandariya"):
        return None
    neighborhood = parts[-1]
    neighborhood = LOCATION_FIXES.get(neighborhood, neighborhood)
    return f"Alexandria / {neighborhood}"


def clean_record(raw: dict) -> dict | None:
    """ينظف record واحد — بيرجع None لو فيه مشكلة"""
    try:
        area = int(raw.get("area(m²)", 0))
        rent = int(raw.get("rent(EGP)", 0))
        loc  = clean_location(str(raw.get("location_en", "")))

        if not loc:
            return None
        if not (AREA_MIN <= area <= AREA_MAX):
            return None
        if not (RENT_MIN <= rent <= RENT_MAX):
            return None

        return {
            "area(m²)":    area,
            "rent(EGP)":   rent,
            "location_en": loc,
        }
    except Exception as e:
        log.debug(f"Clean error: {e} | raw={raw}")
        return None


# ─── Excel Export (نفس شكل rent_commercial3_cleaned.xlsx) ──────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")  # أزرق داكن
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT     = Font(name="Arial", size=10)
EVEN_FILL     = PatternFill("solid", fgColor="D9E1F2")  # أزرق فاتح
THIN_BORDER   = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
COL_WIDTHS    = {"A": 12, "B": 14, "C": 45}   # area | rent | location


def save_to_excel(records: list[dict], cycle: int):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_tag  = datetime.now().strftime("%Y%m%d")
    filename  = f"rent_commercial_alexandria_cycle{cycle}_{date_tag}.xlsx"
    filepath  = os.path.join(OUTPUT_DIR, filename)

    df = (
        pd.DataFrame(records)
        .drop_duplicates(subset=["location_en", "area(m²)", "rent(EGP)"])
        .sort_values("location_en")
        .reset_index(drop=True)
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "rent_commercial_cleaned"

    headers = ["area(m²)", "rent(EGP)", "location_en"]
    for col_idx, header in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = Alignment(horizontal="center", vertical="center")
        cell.border     = THIN_BORDER

    ws.row_dimensions[1].height = 22

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        fill = EVEN_FILL if row_idx % 2 == 0 else PatternFill()
        for col_idx, col_name in enumerate(headers, 1):
            cell           = ws.cell(row=excel_row, column=col_idx, value=row[col_name])
            cell.font      = DATA_FONT
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="center" if col_idx < 3 else "left",
                vertical="center",
            )

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Summary row ──────────────────────────────────────────────────────────
    last_data_row = len(df) + 1
    summary_row   = last_data_row + 2
    ws.cell(row=summary_row, column=1, value="Total Records")
    ws.cell(row=summary_row, column=2, value=f"=COUNTA(B2:B{last_data_row})")
    ws.cell(row=summary_row, column=1).font = Font(name="Arial", bold=True, size=10)

    ws.cell(row=summary_row + 1, column=1, value="Avg Rent (EGP)")
    ws.cell(row=summary_row + 1, column=2, value=f"=AVERAGE(B2:B{last_data_row})")
    ws.cell(row=summary_row + 1, column=1).font = Font(name="Arial", bold=True, size=10)

    ws.cell(row=summary_row + 2, column=1, value="Avg Area (m²)")
    ws.cell(row=summary_row + 2, column=2, value=f"=AVERAGE(A2:A{last_data_row})")
    ws.cell(row=summary_row + 2, column=1).font = Font(name="Arial", bold=True, size=10)

    ws.cell(row=summary_row + 3, column=1, value="Scraped At")
    ws.cell(row=summary_row + 3, column=2, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=summary_row + 3, column=1).font = Font(name="Arial", bold=True, size=10)

    wb.save(filepath)
    log.info(f"Saved {len(df)} records → {filepath}")
    return filepath


# ─── Main Consumer Loop ─────────────────────────────────────────────────────────

def run():
    consumer = KafkaConsumer(
        TOPIC,
        bootstrap_servers=KAFKA_BROKER,
        group_id=GROUP_ID,
        auto_offset_reset="earliest",
        enable_auto_commit=True,
        value_deserializer=lambda m: json.loads(m.decode("utf-8")),
        consumer_timeout_ms=BATCH_TIMEOUT,
    )

    log.info(f"Listening on topic '{TOPIC}' ...")
    cycle      = 1
    buffer     = []

    try:
        for msg in consumer:
            raw     = msg.value
            cleaned = clean_record(raw)
            if cleaned:
                buffer.append(cleaned)

    except StopIteration:
        # consumer_timeout_ms انتهى → البيانات خلصت
        pass
    finally:
        consumer.close()

    if buffer:
        path = save_to_excel(buffer, cycle)
        log.info(f"Cycle {cycle}: {len(buffer)} clean records saved → {path}")
    else:
        log.warning("No valid records received in this cycle.")


if __name__ == "__main__":
    # كل مرة بيتشغل المستهلك (بعد ما الـ Scheduler يشغل الـ Producer)
    # بيستهلك كل الـ events الجديدة ويحفظهم في Excel
    run()
