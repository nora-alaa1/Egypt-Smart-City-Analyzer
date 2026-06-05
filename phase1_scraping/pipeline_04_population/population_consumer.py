"""
Population Data Consumer
=========================
يستقبل من topic:
  population-alexandria  → Population_Alexandria.xlsx

Output columns:
  district | population | population_source | nightlight_intensity | year | latitude | longitude
"""

import json
import logging
import os
from datetime import datetime

import pandas as pd
from kafka import KafkaConsumer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [POP-CONSUMER] %(message)s")
log = logging.getLogger(__name__)

KAFKA_BROKER = os.getenv("KAFKA_BROKER", "localhost:9092")
TOPIC        = "population-alexandria"
GROUP_ID     = "population-cleaner-group"
OUTPUT_DIR   = os.path.join(os.path.dirname(__file__), "output")
TIMEOUT_MS   = 60_000

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
EVEN_FILL    = PatternFill("solid", fgColor="D9E1F2")
THIN         = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
COL_WIDTHS   = {
    "district": 30, "population": 16, "population_source": 18,
    "nightlight_intensity": 22, "year": 8, "latitude": 13, "longitude": 13,
}


def clean_record(raw: dict) -> dict | None:
    try:
        district = str(raw.get("district", "")).strip()
        if not district:
            return None
        pop = raw.get("population")
        if pop is not None:
            pop = int(float(pop))
        return {
            "district":             district,
            "population":           pop,
            "population_source":    raw.get("population_source", "estimate"),
            "nightlight_intensity": raw.get("nightlight_intensity"),
            "year":                 raw.get("year", datetime.now().year),
            "latitude":             raw.get("latitude"),
            "longitude":            raw.get("longitude"),
        }
    except Exception as e:
        log.debug(f"Clean error: {e}")
        return None


def write_excel(df: pd.DataFrame, path: str):
    headers = df.columns.tolist()
    wb = Workbook()
    ws = wb.active
    ws.title = "Population"

    # Title
    ws.merge_cells(f"A1:{chr(64+len(headers))}1")
    tc = ws["A1"]
    tc.value = "Alexandria Population Data — SmartCityAnalyzer"
    tc.font  = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    tc.fill  = PatternFill("solid", fgColor="0D3349")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN
        ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = COL_WIDTHS.get(h, 15)
    ws.row_dimensions[2].height = 20

    # Data
    for ri, row in df.iterrows():
        er   = ri + 3
        fill = EVEN_FILL if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            c = ws.cell(row=er, column=ci, value=val)
            c.font, c.fill, c.border = DATA_FONT, fill, THIN
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center",
            )

    # Summary
    last = len(df) + 2
    sr   = last + 2
    ws.cell(row=sr,   column=1, value="Total Districts").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr,   column=2, value=f"=COUNTA(A3:A{last})")
    ws.cell(row=sr+1, column=1, value="Total Population").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr+1, column=2, value=f"=SUM(B3:B{last})")
    ws.cell(row=sr+2, column=1, value="Avg Nightlight").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr+2, column=2, value=f"=AVERAGE(D3:D{last})")
    ws.cell(row=sr+3, column=1, value="Scraped At").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr+3, column=2, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=sr+4, column=1, value="Source").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr+4, column=2, value="CAPMAS + Google Earth Engine VIIRS")

    wb.save(path)
    log.info(f"Saved {len(df)} rows → {path}")


def run():
    consumer = KafkaConsumer(
        TOPIC,
        bootstrap_servers=KAFKA_BROKER,
        group_id=GROUP_ID,
        auto_offset_reset="earliest",
        enable_auto_commit=True,
        value_deserializer=lambda m: json.loads(m.decode("utf-8")),
        consumer_timeout_ms=TIMEOUT_MS,
    )

    log.info(f"Listening on topic '{TOPIC}' ...")
    buffer = []

    try:
        for msg in consumer:
            r = clean_record(msg.value)
            if r:
                buffer.append(r)
    except StopIteration:
        pass
    finally:
        consumer.close()

    if buffer:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        date_tag = datetime.now().strftime("%Y%m%d")
        df = (
            pd.DataFrame(buffer)
            .drop_duplicates(subset=["district", "year"])
            .sort_values("district")
            .reset_index(drop=True)
        )
        cols = ["district", "population", "population_source",
                "nightlight_intensity", "year", "latitude", "longitude"]
        path = os.path.join(OUTPUT_DIR, f"population_alexandria_{date_tag}.xlsx")
        write_excel(df[cols], path)
        log.info(f"Done — {len(df)} districts saved.")
    else:
        log.warning("No records received.")


if __name__ == "__main__":
    run()
