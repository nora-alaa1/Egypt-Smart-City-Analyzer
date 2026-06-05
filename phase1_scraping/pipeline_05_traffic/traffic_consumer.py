"""
Traffic Data Consumer
======================
يستقبل من topic-ين:
  traffic-nodes-alexandria  → Nodes_Alexandria.xlsx
  traffic-edges-alexandria  → Edges_Alexandria.xlsx

Output:
  Nodes: node_id | latitude | longitude | district | betweenness | street_count | source
  Edges: from_node | to_node | name | length_m | speed_kph | road_type | source
"""

import json
import logging
import os
from datetime import datetime

import pandas as pd
from kafka import KafkaConsumer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [TRAFFIC-CONSUMER] %(message)s")
log = logging.getLogger(__name__)

KAFKA_BROKER = os.getenv("KAFKA_BROKER", "localhost:9092")
TOPICS       = ["traffic-nodes-alexandria", "traffic-edges-alexandria"]
GROUP_ID     = "traffic-cleaner-group"
OUTPUT_DIR   = os.path.join(os.path.dirname(__file__), "output")
TIMEOUT_MS   = 90_000   # أكبر لأن الداتا كبيرة

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT    = Font(name="Arial", size=10)
EVEN_FILL    = PatternFill("solid", fgColor="D9E1F2")
THIN         = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

ALEX_LAT = (30.9, 31.4)
ALEX_LON = (29.5, 30.2)


def clean_node(raw: dict) -> dict | None:
    try:
        lat = float(raw.get("latitude", 0))
        lon = float(raw.get("longitude", 0))
        if not (ALEX_LAT[0] <= lat <= ALEX_LAT[1] and ALEX_LON[0] <= lon <= ALEX_LON[1]):
            return None
        return {
            "node_id":      str(raw.get("node_id", "")),
            "latitude":     round(lat, 6),
            "longitude":    round(lon, 6),
            "district":     raw.get("district"),
            "street_count": raw.get("street_count"),
            "betweenness":  raw.get("betweenness"),
            "source":       raw.get("source", ""),
        }
    except Exception:
        return None


def clean_edge(raw: dict) -> dict | None:
    try:
        length = raw.get("length_m")
        if length is not None:
            length = round(float(length), 1)
        speed = raw.get("speed_kph")
        if speed is not None:
            speed = int(float(speed))
        return {
            "from_node": str(raw.get("from_node", "")),
            "to_node":   str(raw.get("to_node", "")),
            "name":      raw.get("name"),
            "length_m":  length,
            "speed_kph": speed,
            "road_type": raw.get("road_type", "Unclassified"),
            "source":    raw.get("source", ""),
        }
    except Exception:
        return None


def write_excel(df: pd.DataFrame, path: str, sheet_name: str, title: str,
                col_widths: dict):
    headers = df.columns.tolist()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Title row
    ws.merge_cells(f"A1:{chr(64+len(headers))}1")
    tc = ws["A1"]
    tc.value     = title
    tc.font      = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    tc.fill      = PatternFill("solid", fgColor="0D3349")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header row
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = THIN
        ws.column_dimensions[ws.cell(row=2, column=ci).column_letter].width = col_widths.get(h, 16)
    ws.row_dimensions[2].height = 20

    # Data (max 50k rows لأداء Excel)
    df_write = df.head(50000)
    for ri, row in df_write.iterrows():
        er   = ri + 3
        fill = EVEN_FILL if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            c = ws.cell(row=er, column=ci, value=val)
            c.font, c.fill, c.border = DATA_FONT, fill, THIN
            c.alignment = Alignment(
                horizontal="left" if ci <= 2 else "center",
                vertical="center",
            )

    # Summary
    last = len(df_write) + 2
    sr   = last + 2
    ws.cell(row=sr,   column=1, value="Total Records").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr,   column=2, value=f"=COUNTA(A3:A{last})")
    ws.cell(row=sr+1, column=1, value="Scraped At").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr+1, column=2, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=sr+2, column=1, value="Source").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=sr+2, column=2, value="SQLite DB + CSV + OpenStreetMap (Overpass API)")

    if len(df) > 50000:
        ws.cell(row=sr+3, column=1, value="⚠ Note").font = Font(name="Arial", bold=True, color="CC0000", size=10)
        ws.cell(row=sr+3, column=2, value=f"Excel limited to 50,000 rows. Full dataset: {len(df)} rows (see CSV)")

    wb.save(path)
    log.info(f"Saved {min(len(df), 50000)} rows → {path}")

    # حفظ CSV كامل إذا كانت الداتا كبيرة
    if len(df) > 50000:
        csv_path = path.replace(".xlsx", "_full.csv")
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        log.info(f"Full dataset ({len(df)} rows) → {csv_path}")


def run():
    consumer = KafkaConsumer(
        *TOPICS,
        bootstrap_servers=KAFKA_BROKER,
        group_id=GROUP_ID,
        auto_offset_reset="earliest",
        enable_auto_commit=True,
        value_deserializer=lambda m: json.loads(m.decode("utf-8")),
        consumer_timeout_ms=TIMEOUT_MS,
        session_timeout_ms=120_000,
        heartbeat_interval_ms=40_000,
        max_poll_interval_ms=600_000,
        max_poll_records=1000,
        fetch_max_bytes=52428800,
    )

    log.info(f"Listening on topics: {TOPICS}")
    buffers = {"nodes": [], "edges": []}

    try:
        for msg in consumer:
            raw = msg.value
            if msg.topic == "traffic-nodes-alexandria":
                r = clean_node(raw)
                if r:
                    buffers["nodes"].append(r)
            elif msg.topic == "traffic-edges-alexandria":
                r = clean_edge(raw)
                if r:
                    buffers["edges"].append(r)
    except StopIteration:
        pass
    finally:
        consumer.close()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    date_tag = datetime.now().strftime("%Y%m%d")

    # Nodes Excel
    if buffers["nodes"]:
        df_nodes = (
            pd.DataFrame(buffers["nodes"])
            .drop_duplicates(subset=["latitude", "longitude"])
            .sort_values(["district", "latitude"])
            .reset_index(drop=True)
        )
        path = os.path.join(OUTPUT_DIR, f"traffic_nodes_alexandria_{date_tag}.xlsx")
        write_excel(
            df_nodes[["node_id", "latitude", "longitude", "district", "street_count", "betweenness", "source"]],
            path, "Nodes", "Alexandria Traffic Nodes — SmartCityAnalyzer",
            {"node_id": 22, "latitude": 13, "longitude": 13, "district": 25,
             "street_count": 14, "betweenness": 16, "source": 20},
        )

    # Edges Excel
    if buffers["edges"]:
        df_edges = (
            pd.DataFrame(buffers["edges"])
            .drop_duplicates(subset=["from_node", "to_node"])
            .sort_values(["road_type", "name"])
            .reset_index(drop=True)
        )
        path = os.path.join(OUTPUT_DIR, f"traffic_edges_alexandria_{date_tag}.xlsx")
        write_excel(
            df_edges[["from_node", "to_node", "name", "length_m", "speed_kph", "road_type", "source"]],
            path, "Edges", "Alexandria Traffic Edges (Roads) — SmartCityAnalyzer",
            {"from_node": 22, "to_node": 22, "name": 35, "length_m": 12,
             "speed_kph": 12, "road_type": 18, "source": 20},
        )

    log.info(
        f"Done — nodes:{len(buffers['nodes'])} edges:{len(buffers['edges'])}"
    )


if __name__ == "__main__":
    run()
