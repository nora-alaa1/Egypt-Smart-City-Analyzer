"""
test_pipeline.py — 50 tests — بدون Kafka/internet
"""
import re, json, sys
from collections import deque
from bs4 import BeautifulSoup
from datetime import datetime, timezone
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# BASE is relative to the repo root — 4 levels up from this file
import pathlib as _pl
BASE      = str(_pl.Path(__file__).parents[3] / "data")
CAFES_DF  = pd.read_excel(f"{BASE}/cafes_data/cafes.xlsx")
GYMS_DF   = pd.read_excel(f"{BASE}/gym_data/gyms.xlsx")
PHARMA_DF = pd.read_csv(f"{BASE}/Pharmacies_Data/processed/pharmacies NORMALIZED  (1).csv")

PASS, FAIL = 0, 0
def check(cond, ok_msg, fail_msg=None):
    global PASS, FAIL
    if cond: PASS += 1; print(f"  ✅ {ok_msg}")
    else:    FAIL += 1; print(f"  ❌ {fail_msg or ok_msg}")

# ── Shared logic ────────────────────────────────────────────────────────────
ALEX_LAT = (30.9, 31.4)
ALEX_LON = (29.5, 30.2)
RE_PHONE = re.compile(r"[\d\+\-\(\)\s\.]{7,20}")

def extract_coords(el):
    if el["type"] == "node": return el.get("lat"), el.get("lon")
    c = el.get("center", {}); return c.get("lat"), c.get("lon")

def parse_osm_element(el, category):
    tags = el.get("tags", {})
    name = (tags.get("name") or tags.get("name:ar") or
            tags.get("name:en") or tags.get("brand") or None)
    lat, lon = extract_coords(el)
    if lat is None or lon is None: return None
    return {"name": name, "latitude": round(float(lat), 6),
            "longitude": round(float(lon), 6), "osm_id": el.get("id"),
            "category": category}

def clean_osm_record(raw):
    try:
        lat  = float(raw.get("latitude", 0))
        lon  = float(raw.get("longitude", 0))
        name = raw.get("name")
        if not (ALEX_LAT[0] <= lat <= ALEX_LAT[1]): return None
        if not (ALEX_LON[0] <= lon <= ALEX_LON[1]): return None
        return {"name": name, "latitude": round(lat, 6), "longitude": round(lon, 6)}
    except: return None

def clean_pharmacy_record(raw):
    try:
        name    = str(raw.get("name",    "")).strip()
        address = str(raw.get("address", "")).strip()
        phone   = raw.get("phone")
        url     = str(raw.get("web_scraper_start_url", "")).strip()
        src     = str(raw.get("source_file", "")).strip()
        if not name or not address: return None
        if phone:
            phone_str = str(phone).strip()
            digits    = re.sub(r"[^\d]", "", phone_str)
            if len(digits) < 5:
                phone = None          # short → nullify, don't drop record
            else:
                try:
                    # حاول تحويل لـ float — لو فشل (زي +20.12.x) احتفظ كـ string
                    if phone_str.replace(",","").replace(".","",1).replace("+","",1).replace("-","",1).lstrip("-").isdigit():
                        phone = float(phone_str.replace(",",""))
                    else:
                        phone = phone_str  # international format → keep as string
                except: phone = phone_str
        return {"name": name, "address": address, "phone": phone,
                "source_file": src, "web_scraper_start_url": url}
    except: return None

def parse_pharmacy_card(card):
    try:
        name_el = card.select_one(
            "h2.listing-title,h3.listing-title,.business-name,h2,h3")
        name = name_el.get_text(strip=True) if name_el else None
        if not name: return None
        addr_el = card.select_one(
            ".listing-address,.address,[class*='address'],p.location")
        if addr_el:
            address = addr_el.get_text(strip=True)
        else:
            all_text = [p.get_text(strip=True) for p in card.select("p,span,div")]
            cands = [t for t in all_text
                     if any(kw in t.lower()
                            for kw in ["st.", "rd.", "ave", "street", "road", "el ", " - "])
                     and len(t) > 10]
            address = cands[0] if cands else ""
        if not address: return None          # ← لازم يكون موجود
        phone_el = card.select_one(
            ".listing-phone,.phone,[class*='phone'],[href^='tel:']")
        if phone_el:
            phone_raw = (phone_el.get("href", "").replace("tel:", "")
                         or phone_el.get_text(strip=True))
        else:
            m = RE_PHONE.search(card.get_text(" "))
            phone_raw = m.group().strip() if m else None
        return {"name": name, "address": address, "phone": phone_raw}
    except: return None


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST 1 — OSM parse_osm_element")
print("═"*62)

node_cafe    = {"type":"node","id":1,"lat":31.2156,"lon":29.9553,
                "tags":{"amenity":"cafe","name":"Carlos Café"}}
node_no_name = {"type":"node","id":2,"lat":31.219, "lon":29.948,
                "tags":{"amenity":"cafe"}}
node_arabic  = {"type":"node","id":3,"lat":31.226, "lon":29.946,
                "tags":{"name:ar":"كافيه السلام","amenity":"cafe"}}
node_way     = {"type":"way", "id":4,
                "center":{"lat":31.226,"lon":29.946},
                "tags":{"name":"Buzz Cafe"}}
node_no_coord= {"type":"node","id":5,"tags":{"name":"Ghost"}}

r1 = parse_osm_element(node_cafe,    "cafes")
r2 = parse_osm_element(node_no_name, "cafes")
r3 = parse_osm_element(node_arabic,  "cafes")
r4 = parse_osm_element(node_way,     "cafes")
r5 = parse_osm_element(node_no_coord,"cafes")

check(r1 is not None,                         "Node parsed successfully")
check(r1["name"] == "Carlos Café",            f"EN name: '{r1['name']}'")
check(r1["latitude"]  == 31.2156,             f"Lat: {r1['latitude']}")
check(r1["longitude"] == 29.9553,             f"Lon: {r1['longitude']}")
check(r2 is not None and r2["name"] is None,  "None name accepted (like cafes.xlsx 88 nulls)")
check(r3 is not None and r3["name"]=="كافيه السلام", f"Arabic name: '{r3['name']}'")
check(r4 is not None and r4["latitude"]==31.226,"Way center coords")
check(r5 is None,                             "No coords → filtered")


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST 2 — OSM clean_osm_record: bbox filter")
print("═"*62)

osm_cases = [
    ({"latitude":31.2151,"longitude":29.9608,"name":"Sidi Gaber"}, True,  "Valid Alexandria coord"),
    ({"latitude":33.0,   "longitude":31.0,   "name":"Cairo"},      False, "Outside: Cairo"),
    ({"latitude":31.2,   "longitude":28.0,   "name":"West"},       False, "Outside: lon too low"),
    ({"latitude":30.85,  "longitude":29.8,   "name":"South"},      False, "Outside: lat too low"),
    ({"latitude":31.2,   "longitude":29.9,   "name":None},         True,  "None name accepted"),
    ({"latitude":31.261512,"longitude":29.984544,"name":"Hammam"}, True,  "Real gym: Hammam Gym"),
    ({"latitude":31.226593,"longitude":29.945837,"name":"نادي الفتح"},True,"Real gym: نادي الفتح"),
    ({"latitude":31.0,   "longitude":30.3,   "name":"East"},       False, "Outside: lon too high"),
]
for raw, should, label in osm_cases:
    r = clean_osm_record(raw)
    check((r is not None) == should, label,
          f"WRONG: {label} — got {'PASS' if r else 'FILTER'}")

# كل الداتا الحقيقية لازم تعدي
gym_pass  = sum(1 for _, row in GYMS_DF.iterrows()
                if clean_osm_record({"latitude":row["latitude"],"longitude":row["longitude"],"name":row["name"]}))
cafe_pass = sum(1 for _, row in CAFES_DF.iterrows()
                if clean_osm_record({"latitude":row["latitude"],"longitude":row["longitude"],"name":row.get("name")}))

check(gym_pass  == len(GYMS_DF),  f"All {len(GYMS_DF)} real gyms  pass bbox ({gym_pass}/{len(GYMS_DF)})")
check(cafe_pass == len(CAFES_DF), f"All {len(CAFES_DF)} real cafes pass bbox ({cafe_pass}/{len(CAFES_DF)})")


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST 3 — Pharmacy: HTML parsing")
print("═"*62)

PHARMA_HTML = """<html><body>
<div class="listing-item">
  <h2 class="listing-title">Dr. Osama El Tayeby Pharmacies El Seyouf Branch</h2>
  <p class="listing-address">amin hasouna st., intersection of mostafa kamel st. - inside city light mall, shop 9</p>
  <p class="listing-phone">35301103</p>
</div>
<div class="listing-item">
  <h2 class="listing-title">El Beisy Pharmacies Semouha Branch</h2>
  <p class="listing-address">7 bahaa el din el ghatwary st., off fawzy moaz st. - near national bank of egypt</p>
  <p class="listing-phone">34040235</p>
</div>
<div class="listing-item">
  <h2 class="listing-title">Balbaa Pharmacies Miami Branch</h2>
  <p class="listing-address">189 khaled ibn el walid st. - near to el montazah cinema</p>
  <a href="tel:35576737" class="listing-phone">35576737</a>
</div>
<div class="listing-item">
  <h2 class="listing-title">No Address Pharmacy</h2>
</div>
</body></html>"""

soup  = BeautifulSoup(PHARMA_HTML, "lxml")
cards = soup.select("div.listing-item")
recs  = [parse_pharmacy_card(c) for c in cards]
valid = [r for r in recs if r]

check(len(cards) == 4,  f"Found {len(cards)} pharmacy cards")
check(len(valid) == 3,  f"Parsed {len(valid)}/4 (1 filtered: empty address)")
check(recs[3] is None,  "Card 4 (no address) correctly returns None")
check(valid[0]["name"] == "Dr. Osama El Tayeby Pharmacies El Seyouf Branch",
      f"Name: '{valid[0]['name'][:45]}'")
check("amin hasouna" in valid[0]["address"].lower(), "Address from .listing-address")
check(valid[0]["phone"] == "35301103",               "Phone from <p class='listing-phone'>")
check(valid[2]["phone"] == "35576737",               "Phone from <a href='tel:'>")


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST 4 — Pharmacy: clean_pharmacy_record")
print("═"*62)

ph_cases = [
    # (raw, should_pass, phone_should_be_null, label)
    ({"name":"Balbaa Pharmacies Miami Branch",
      "address":"189 khaled ibn el walid st.",
      "phone":"35576737",
      "source_file":"egyfinder page 2",
      "web_scraper_start_url":"https://egyfinder.net/categories/en/pharmacies/alexandria?p=2"},
     True, False, "Valid full record"),

    ({"name":"","address":"some st.","phone":"35576737",
      "source_file":"p","web_scraper_start_url":"url"},
     False, False, "Empty name → record filtered"),

    ({"name":"Test","address":"","phone":"35576737",
      "source_file":"p","web_scraper_start_url":"url"},
     False, False, "Empty address → record filtered"),

    ({"name":"Test","address":"some st.","phone":None,
      "source_file":"p","web_scraper_start_url":"url"},
     True, True, "None phone → record passes, phone=None"),

    ({"name":"Test","address":"some st.","phone":"123",
      "source_file":"p","web_scraper_start_url":"url"},
     True, True, "Short phone (<5 digits) → record passes, phone=None"),

    ({"name":"Test","address":"some st.","phone":"+20.12.2377.1659",
      "source_file":"p","web_scraper_start_url":"url"},
     True, False, "International phone format accepted"),
]

for raw, should_pass, phone_null, label in ph_cases:
    r = clean_pharmacy_record(raw)
    check((r is not None) == should_pass, label,
          f"WRONG record filter: {label}")
    if should_pass and r is not None:
        if phone_null:
            check(r["phone"] is None, f"  phone=None correctly: {label}",
                  f"  WRONG phone not nulled: got {r['phone']}")
        else:
            check(r["phone"] is not None, f"  phone intact: {label}",
                  f"  WRONG phone dropped: {label}")

# كل الداتا الحقيقية لازم تعدي
pharma_pass = sum(1 for _, row in PHARMA_DF.iterrows()
                  if clean_pharmacy_record({
                      "name":     row["name"],
                      "address":  row["address"],
                      "phone":    row["phone"],
                      "source_file": row["source_file"],
                      "web_scraper_start_url": row["web_scraper_start_url"],
                  }))
check(pharma_pass == len(PHARMA_DF),
      f"All {len(PHARMA_DF)} real pharmacy records pass cleaning ({pharma_pass}/{len(PHARMA_DF)})")


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST 5 — Mock Kafka: end-to-end roundtrip (real data)")
print("═"*62)

qs = {
    "osm-cafes-alexandria": deque(),
    "osm-gyms-alexandria":  deque(),
    "pharmacy-alexandria":  deque(),
}

class MockProducer:
    def send(self, topic, value):
        qs[topic].append(json.dumps(value, ensure_ascii=False).encode())
    def flush(self): pass
    def close(self): pass

class MockMsg:
    def __init__(self, topic, raw):
        self.topic = topic
        self.value = json.loads(raw.decode())

class MockConsumer:
    def __iter__(self):
        for t, q in qs.items():
            while q: yield MockMsg(t, q.popleft())

mp = MockProducer()

# Send ALL real gyms
for _, row in GYMS_DF.iterrows():
    mp.send("osm-gyms-alexandria", {
        "name": row["name"], "latitude": row["latitude"],
        "longitude": row["longitude"], "category": "gyms",
        "scraped_at": datetime.now(timezone.utc).isoformat(),
    })

# Send ALL real cafes
for _, row in CAFES_DF.iterrows():
    mp.send("osm-cafes-alexandria", {
        "name": None if pd.isna(row.get("name")) else row.get("name"), "latitude": row["latitude"],
        "longitude": row["longitude"], "category": "cafes",
        "scraped_at": datetime.now(timezone.utc).isoformat(),
    })

# Send ALL real pharmacies
for _, row in PHARMA_DF.iterrows():
    mp.send("pharmacy-alexandria", {
        "name":    row["name"],    "address": row["address"],
        "phone":   row["phone"],   "source_file": row["source_file"],
        "web_scraper_start_url": row["web_scraper_start_url"],
        "scraped_at": datetime.now(timezone.utc).isoformat(),
    })

total_sent = sum(len(q) for q in qs.values())
check(total_sent == len(GYMS_DF) + len(CAFES_DF) + len(PHARMA_DF),
      f"Sent {total_sent} = {len(GYMS_DF)}+{len(CAFES_DF)}+{len(PHARMA_DF)} messages")

bufs = {"cafes": [], "gyms": [], "pharmacies": []}
for msg in MockConsumer():
    if   msg.topic == "osm-cafes-alexandria":
        r = clean_osm_record(msg.value)
        if r: bufs["cafes"].append(r)
    elif msg.topic == "osm-gyms-alexandria":
        r = clean_osm_record(msg.value)
        if r: bufs["gyms"].append(r)
    elif msg.topic == "pharmacy-alexandria":
        r = clean_pharmacy_record(msg.value)
        if r: bufs["pharmacies"].append(r)

check(len(bufs["gyms"])   == len(GYMS_DF),   f"Gyms roundtrip: {len(bufs['gyms'])}/{len(GYMS_DF)}")
check(len(bufs["cafes"])  == len(CAFES_DF),  f"Cafes roundtrip: {len(bufs['cafes'])}/{len(CAFES_DF)}")
check(len(bufs["pharmacies"]) == len(PHARMA_DF),
      f"Pharmacies roundtrip: {len(bufs['pharmacies'])}/{len(PHARMA_DF)}")

# No message loss
check(all(q == deque() for q in qs.values()), "All topics empty after consume (no message loss)")

# Data integrity checks
check(bufs["gyms"][0]["name"] == GYMS_DF.iloc[0]["name"],
      f"Gym[0] name intact: '{bufs['gyms'][0]['name']}'")
check(abs(bufs["gyms"][0]["latitude"] - float(GYMS_DF.iloc[0]["latitude"])) < 0.00001,
      f"Gym[0] lat intact: {bufs['gyms'][0]['latitude']}")
check(bufs["pharmacies"][0]["name"] == PHARMA_DF.iloc[0]["name"],
      f"Pharmacy[0] name intact: '{bufs['pharmacies'][0]['name']}'")
check(bufs["pharmacies"][0]["address"] == PHARMA_DF.iloc[0]["address"],
      "Pharmacy[0] address intact")

# Null handling
cafe_nulls_in  = CAFES_DF["name"].isna().sum()
cafe_nulls_out = sum(1 for r in bufs["cafes"] if r["name"] is None)
check(cafe_nulls_in == cafe_nulls_out,
      f"Null names preserved: {cafe_nulls_out} nulls (matches original {cafe_nulls_in})")


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST 6 — Output format: Excel (gyms/cafes) + CSV (pharmacies)")
print("═"*62)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)
EVEN_FILL   = PatternFill("solid", fgColor="D9E1F2")
THIN        = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"),  bottom=Side(style="thin"))

def write_xl(df, path, sheet):
    headers = df.columns.tolist()
    wb = Workbook(); ws = wb.active; ws.title = sheet
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN
    for ri, row in df.iterrows():
        er = ri + 2
        fill = EVEN_FILL if ri % 2 == 0 else PatternFill()
        for ci, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val): val = None
            c = ws.cell(row=er, column=ci, value=val)
            c.font, c.fill, c.border = DATA_FONT, fill, THIN
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center")
    wb.save(path)

# ── Gyms Excel ───────────────────────────────────────────────────────────
df_g = pd.DataFrame(bufs["gyms"])[["name", "latitude", "longitude"]]
write_xl(df_g, "/tmp/out_gyms.xlsx", "gyms")
wg = load_workbook("/tmp/out_gyms.xlsx").active

check(wg.title          == "gyms",      "Gyms: sheet name = 'gyms'")
check(wg["A1"].value    == "name",      "Gyms: A1 = 'name'")
check(wg["B1"].value    == "latitude",  "Gyms: B1 = 'latitude'")
check(wg["C1"].value    == "longitude", "Gyms: C1 = 'longitude'")
check(wg["A1"].font.bold,               "Gyms: header bold")
check("1F4E79" in wg["A1"].fill.fgColor.rgb, "Gyms: header navy fill")
check(wg.max_row        == len(df_g)+1, f"Gyms: {wg.max_row-1} data rows = {len(df_g)}")
check(wg["A2"].value    == GYMS_DF.iloc[0]["name"],
      f"Gyms[0] name: '{wg['A2'].value}'")
check(abs(wg["B2"].value - float(GYMS_DF.iloc[0]["latitude"])) < 0.00001,
      f"Gyms[0] lat: {wg['B2'].value}")
check(abs(wg["C2"].value - float(GYMS_DF.iloc[0]["longitude"])) < 0.00001,
      f"Gyms[0] lon: {wg['C2'].value}")

# ── Cafes Excel ───────────────────────────────────────────────────────────
df_c = pd.DataFrame(bufs["cafes"])[["name", "latitude", "longitude"]]
write_xl(df_c, "/tmp/out_cafes.xlsx", "cafes")
wc = load_workbook("/tmp/out_cafes.xlsx").active

check(wc.title       == "cafes",     "Cafes: sheet name = 'cafes'")
check(wc["A1"].value == "name",      "Cafes: A1 = 'name'")
check(wc["B1"].value == "latitude",  "Cafes: B1 = 'latitude'")
check(wc["C1"].value == "longitude", "Cafes: C1 = 'longitude'")
check(wc.max_row     == len(df_c)+1, f"Cafes: {wc.max_row-1} data rows = {len(df_c)}")

# null names stored as None not "None"
null_cells = [wc.cell(row=r, column=1).value
              for r in range(2, wc.max_row+1)
              if wc.cell(row=r, column=1).value is None]
check(len(null_cells) == cafe_nulls_out,
      f"Cafes: {len(null_cells)} null names stored as None (not string 'None')")

# ── Pharmacies CSV ────────────────────────────────────────────────────────
df_p = pd.DataFrame(bufs["pharmacies"])[
    ["name", "address", "phone", "source_file", "web_scraper_start_url"]]
df_p.to_csv("/tmp/out_pharmacies.csv", index=False, encoding="utf-8-sig")
df_back = pd.read_csv("/tmp/out_pharmacies.csv")

check(list(df_back.columns) == ["name","address","phone","source_file","web_scraper_start_url"],
      "Pharmacies: columns match target exactly")
check(len(df_back)           == len(PHARMA_DF),
      f"Pharmacies: {len(df_back)} rows = {len(PHARMA_DF)}")
check(df_back.iloc[0]["name"]    == PHARMA_DF.iloc[0]["name"],
      f"Pharmacy[0] name: '{df_back.iloc[0]['name'][:40]}'")
check(df_back.iloc[0]["address"] == PHARMA_DF.iloc[0]["address"],
      "Pharmacy[0] address matches")
check("egyfinder" in str(df_back.iloc[0]["web_scraper_start_url"]),
      "Pharmacy[0] URL contains 'egyfinder'")


# ════════════════════════════════════════════════════════════════════════════
print("\n" + "═"*62)
print("  TEST RESULTS")
print("═"*62)
total = PASS + FAIL
print(f"\n  Passed : {PASS}/{total}")
print(f"  Failed : {FAIL}/{total}")
if FAIL == 0:
    print("\n  🎉 ALL TESTS PASSED — ready for docker-compose up")
    print(f"\n  📊 Output format verified against real data:")
    print(f"     cafes.xlsx    → {len(CAFES_DF):>3} rows | name(nullable), latitude, longitude")
    print(f"     gyms.xlsx     → {len(GYMS_DF):>3} rows | name, latitude, longitude")
    print(f"     pharmacies    → {len(PHARMA_DF):>3} rows | name, address, phone, source_file, url")
else:
    print("\n  ⚠️  Some tests failed")
    sys.exit(1)
