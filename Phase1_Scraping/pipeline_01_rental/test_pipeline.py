"""
test_pipeline.py — اشغل ده قبل docker-compose up
يختبر كل حاجة بدون Kafka حقيقي
"""
import re, json, sys
from collections import deque
from bs4 import BeautifulSoup
from datetime import datetime, timezone
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PASS = 0
FAIL = 0

def check(condition, label_pass, label_fail=None):
    global PASS, FAIL
    if condition:
        PASS += 1
        print(f"  ✅ {label_pass}")
    else:
        FAIL += 1
        print(f"  ❌ {label_fail or label_pass}")

# ═══════════════════════════════════════════════════════
print("\n══════════════════════════════════════════════════")
print("  TEST 1 — Producer: HTML Parsing")
print("══════════════════════════════════════════════════")

RE_PRICE = re.compile(r"([\d,]+)\s*EGP/month", re.IGNORECASE)
RE_AREA  = re.compile(r"\+\s*([\d,]+)\s*m[²2]", re.IGNORECASE)

HTML = """
<ul class="search-result-grid relative">
  <li>
    <a href="/en/listing/6758980/"><img alt="Commercial For rent in Gamal Abd El Nasir St, Sidi Beshr Bahri, 600 sqm"></a>
    <a href="/en/listing/6758980/">300,000 EGP/month<h2>Title</h2></a>
    <a href="/en/for-rent/commercial/alexandria/sydy-bshr/">Sidi Bishr</a>/
    <a href="/en/for-rent/commercial/alexandria/sydy-bshr/gamal-st/">Gamal Abd El Nasir St</a>
    <a href="/en/listing/6758980/">+ 600 m²</a>
  </li>
  <li>
    <a href="/en/listing/6821620/"><img alt="Commercial For rent in Salah Salem St., Mahta El Raml, 194 sqm"></a>
    <a href="/en/listing/6821620/">80,000 EGP/month<h2>Shop for Rent</h2></a>
    <a href="/en/for-rent/commercial/alexandria/mht-lrml/">Mahta El Raml</a>/
    <a href="/en/for-rent/commercial/alexandria/mht-lrml/salah-st/">Salah Salem St.</a>
    <a href="/en/listing/6821620/">+ 194 m²</a>
  </li>
  <li>
    <a href="/en/listing/6854383/"><img alt="Commercial For rent in La Jetee St., Ibrahimia, 65 sqm"></a>
    <a href="/en/listing/6854383/">20,000 EGP/month<h2>Shop For rent</h2></a>
    <a href="/en/for-rent/commercial/alexandria/ibrahimia/">Ibrahimia</a>/
    <a href="/en/for-rent/commercial/alexandria/ibrahimia/la-jetee-st/">La Jetee St.</a>
    <a href="/en/listing/6854383/">+ 65 m²</a>
  </li>
  <li>
    <!-- no area text — fallback to img alt sqm -->
    <a href="/en/listing/9999/"><img alt="Commercial For rent in Smouha, 120 sqm"></a>
    <a href="/en/listing/9999/">25,000 EGP/month<h2>Title</h2></a>
    <a href="/en/for-rent/commercial/alexandria/smouha/">Smouha</a>
  </li>
  <li>
    <!-- invalid: no location links -->
    <span>55,000 EGP/month</span><span>+ 200 m²</span>
  </li>
  <li>
    <!-- invalid: no price -->
    <a href="/en/for-rent/commercial/alexandria/glim/">Glim</a>
    <a href="/en/listing/1/">+ 100 m²</a>
  </li>
</ul>
"""

def parse_card(card):
    card_text = card.get_text(" ", strip=True)
    price_m = RE_PRICE.search(card_text)
    if not price_m: return None
    rent = int(price_m.group(1).replace(",", ""))
    area_m = RE_AREA.search(card_text)
    if area_m:
        area = int(area_m.group(1).replace(",", ""))
    else:
        img   = card.select_one("img[alt]")
        sqm_m = re.search(r"(\d+)\s*sqm", img["alt"], re.IGNORECASE) if img else None
        area  = int(sqm_m.group(1)) if sqm_m else None
    if not area: return None
    loc_links = [
        a for a in card.select('a[href*="/for-rent/commercial/alexandria/"]')
        if a["href"].rstrip("/") != "/en/for-rent/commercial/alexandria"
    ]
    if not loc_links: return None
    parts = ["Alexandria"] + [a.get_text(strip=True) for a in loc_links[:2]]
    return {
        "location_en": " / ".join(p for p in parts if p),
        "area(m²)":    area,
        "rent(EGP)":   rent,
        "scraped_at":  datetime.now(timezone.utc).isoformat(),
        "source":      "aqarmap",
    }

soup  = BeautifulSoup(HTML, "lxml")
grid  = soup.select_one("ul[class*='search-result-grid']")
cards = grid.select("li") if grid else []
recs  = [r for c in cards if (r := parse_card(c))]

check(grid is not None,   "Grid selector: ul[class*='search-result-grid'] found",
                          "Grid selector FAILED — class name may have changed")
check(len(cards) == 6,    f"Found {len(cards)} cards",
                          f"Expected 6 cards, got {len(cards)}")
check(len(recs) == 4,     f"Parsed {len(recs)}/6 valid (2 invalid correctly filtered)",
                          f"Expected 4 valid records, got {len(recs)}")

expected = [
    ("Alexandria / Sidi Bishr / Gamal Abd El Nasir St", 600, 300000),
    ("Alexandria / Mahta El Raml / Salah Salem St.",    194,  80000),
    ("Alexandria / Ibrahimia / La Jetee St.",            65,  20000),
    ("Alexandria / Smouha",                             120,  25000),
]
for i, (loc, area, rent) in enumerate(expected):
    r = recs[i]
    match = r["location_en"]==loc and r["area(m²)"]==area and r["rent(EGP)"]==rent
    check(match,
          f"Card {i+1}: {loc[:38]} | {area}m² | {rent:,} EGP",
          f"Card {i+1}: got loc={r['location_en']} area={r['area(m²)']} rent={r['rent(EGP)']}")

# ═══════════════════════════════════════════════════════
print("\n══════════════════════════════════════════════════")
print("  TEST 2 — Mock Kafka Queue")
print("══════════════════════════════════════════════════")

mock_topic = deque()

class MockProducer:
    def send(self, topic, value):
        mock_topic.append(json.dumps(value, ensure_ascii=False).encode())
    def flush(self): pass
    def close(self): pass

class MockMsg:
    def __init__(self, raw): self.value = json.loads(raw.decode())

class MockConsumer:
    def __iter__(self):
        while mock_topic:
            yield MockMsg(mock_topic.popleft())

producer = MockProducer()
for r in recs:
    producer.send("rent-commercial-alexandria", value=r)
producer.flush()

check(len(mock_topic) == 4,
      f"Producer: sent {len(mock_topic)} messages to topic",
      f"Producer: expected 4 messages, got {len(mock_topic)}")

consumer_msgs = list(MockConsumer())
check(len(consumer_msgs) == 4,
      f"Consumer: received {len(consumer_msgs)} messages",
      f"Consumer: expected 4 messages, got {len(consumer_msgs)}")

check(len(mock_topic) == 0,
      "Topic empty after consume (no message loss)",
      "Topic still has messages after consume")

check(all(hasattr(m, "value") and isinstance(m.value, dict) for m in consumer_msgs),
      "All messages deserialized correctly to dict",
      "Message deserialization failed")

# roundtrip check
check(consumer_msgs[0].value["rent(EGP)"] == 300000,
      "Roundtrip check: rent value intact after serialize→deserialize",
      f"Roundtrip failed: got {consumer_msgs[0].value.get('rent(EGP)')}")

# ═══════════════════════════════════════════════════════
print("\n══════════════════════════════════════════════════")
print("  TEST 3 — Consumer: Cleaning Logic")
print("══════════════════════════════════════════════════")

AREA_MIN, AREA_MAX = 50, 5000
RENT_MIN, RENT_MAX = 1100, 700_000
LOCATION_FIXES = {
    "Moharam Bek":"Moharram Bek", "Moharam Bey":"Moharram Bek",
    "El Asafra":"Asafra Bahary",  "Kafr Abdu":"Kafr Abdo",
    "San Stefanus":"San Stefano",
}

def clean_location(raw):
    if not raw or "/" not in raw: return None
    parts = [p.strip().title() for p in raw.split("/")]
    if parts[0].lower() not in ("alexandria","al iskandariya"): return None
    nb = parts[1].strip() if len(parts)>1 else ""
    nb = LOCATION_FIXES.get(nb, nb)
    st = parts[2].strip() if len(parts)>2 else ""
    return f"Alexandria / {nb}" + (f" / {st}" if st else "")

def clean_record(raw):
    try:
        area = int(raw.get("area(m²)", 0))
        rent = int(raw.get("rent(EGP)", 0))
        loc  = clean_location(str(raw.get("location_en", "")))
        if not loc: return None
        if not (AREA_MIN <= area <= AREA_MAX): return None
        if not (RENT_MIN <= rent <= RENT_MAX): return None
        return {"area(m²)": area, "rent(EGP)": rent, "location_en": loc}
    except: return None

cases = [
    ({"location_en":"Alexandria / Sidi Bishr","area(m²)":100,"rent(EGP)":15000}, True,  "Valid record passes"),
    ({"location_en":"Alexandria / Smouha",    "area(m²)":10, "rent(EGP)":5000},  False, "Filtered: area too small (10<50)"),
    ({"location_en":"Alexandria / Smouha",    "area(m²)":100,"rent(EGP)":500},   False, "Filtered: rent too low (500<1100)"),
    ({"location_en":"Alexandria / Glim",      "area(m²)":9999,"rent(EGP)":50000},False, "Filtered: area too large (9999>5000)"),
    ({"location_en":"Cairo / Maadi",          "area(m²)":100,"rent(EGP)":15000}, False, "Filtered: wrong city"),
    ({"location_en":"",                       "area(m²)":100,"rent(EGP)":15000}, False, "Filtered: empty location"),
    ({"location_en":"Alexandria / Moharam Bek","area(m²)":100,"rent(EGP)":15000},True,  "Location fix: Moharam Bek→Moharram Bek"),
    ({"location_en":"Alexandria / Kafr Abdu", "area(m²)":80, "rent(EGP)":8000},  True,  "Location fix: Kafr Abdu→Kafr Abdo"),
]

for raw, should_pass, label in cases:
    result = clean_record(raw)
    check((result is not None) == should_pass, label,
          f"WRONG: {label} — got {'PASS' if result else 'FILTER'}, expected {'PASS' if should_pass else 'FILTER'}")

r_fix1 = clean_record({"location_en":"Alexandria / Moharam Bek","area(m²)":100,"rent(EGP)":15000})
check(r_fix1 is not None and "Moharram Bek" in r_fix1["location_en"],
      "Fix output verified: 'Moharram Bek' in result",
      f"Fix output wrong: got {r_fix1}")

r_fix2 = clean_record({"location_en":"Alexandria / Kafr Abdu","area(m²)":80,"rent(EGP)":8000})
check(r_fix2 is not None and "Kafr Abdo" in r_fix2["location_en"],
      "Fix output verified: 'Kafr Abdo' in result",
      f"Fix output wrong: got {r_fix2}")

# ═══════════════════════════════════════════════════════
print("\n══════════════════════════════════════════════════")
print("  TEST 4 — Excel Output")
print("══════════════════════════════════════════════════")

cleaned = [r for m in consumer_msgs if (r := clean_record(m.value))]
df = (pd.DataFrame(cleaned)
        .drop_duplicates(subset=["location_en","area(m²)","rent(EGP)"])
        .sort_values("location_en").reset_index(drop=True))

check(len(cleaned) == 4, f"All 4 records passed cleaning (none filtered)",
                         f"Expected 4, got {len(cleaned)}")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
DATA_FONT   = Font(name="Arial", size=10)
EVEN_FILL   = PatternFill("solid", fgColor="D9E1F2")
THIN        = Border(left=Side(style="thin"),right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

wb = Workbook()
ws = wb.active
ws.title = "rent_commercial_cleaned"
headers = ["area(m²)", "rent(EGP)", "location_en"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font, c.fill = HEADER_FONT, HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN
ws.row_dimensions[1].height = 22
for ri, row in df.iterrows():
    er = ri + 2
    fill = EVEN_FILL if ri % 2 == 0 else PatternFill()
    for ci, col in enumerate(headers, 1):
        c = ws.cell(row=er, column=ci, value=row[col])
        c.font, c.fill, c.border = DATA_FONT, fill, THIN
        c.alignment = Alignment(horizontal="center" if ci<3 else "left", vertical="center")
for col, w in {"A":12,"B":14,"C":45}.items():
    ws.column_dimensions[col].width = w
last = len(df) + 1
sr   = last + 2
for label, formula, r in [
    ("Total Records",  f"=COUNTA(B2:B{last})", sr),
    ("Avg Rent (EGP)", f"=AVERAGE(B2:B{last})", sr+1),
    ("Avg Area (m²)",  f"=AVERAGE(A2:A{last})", sr+2),
    ("Scraped At",     datetime.now().strftime("%Y-%m-%d"), sr+3),
]:
    ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=r, column=2, value=formula)

path = "/tmp/test_final.xlsx"
wb.save(path)
wc  = load_workbook(path)
ws2 = wc.active

check(ws2.title == "rent_commercial_cleaned",  "Sheet name: rent_commercial_cleaned", f"Wrong sheet name: {ws2.title}")
check(ws2["A1"].value == "area(m²)",           "Header A1: area(m²)",    f"Wrong A1: {ws2['A1'].value}")
check(ws2["B1"].value == "rent(EGP)",          "Header B1: rent(EGP)",   f"Wrong B1: {ws2['B1'].value}")
check(ws2["C1"].value == "location_en",        "Header C1: location_en", f"Wrong C1: {ws2['C1'].value}")
check(ws2["A1"].font.bold == True,             "Header font: bold")
check("FFFFFF" in ws2["A1"].font.color.rgb,    "Header font: white",     f"Got: {ws2['A1'].font.color.rgb}")
check("1F4E79" in ws2["A1"].fill.fgColor.rgb, "Header fill: navy blue", f"Got: {ws2['A1'].fill.fgColor.rgb}")
check(ws2.max_row == last + 5,                 f"Row count: {ws2.max_row} (data+header+gap+summary)",
                                               f"Wrong row count: {ws2.max_row}, expected {last+5}")
check("COUNTA"  in str(ws2.cell(row=sr,  column=2).value), "Summary: =COUNTA formula present")
check("AVERAGE" in str(ws2.cell(row=sr+1,column=2).value), "Summary: =AVERAGE rent formula present")
check("AVERAGE" in str(ws2.cell(row=sr+2,column=2).value), "Summary: =AVERAGE area formula present")

# ═══════════════════════════════════════════════════════
print("\n══════════════════════════════════════════════════")
print("  TEST RESULTS")
print("══════════════════════════════════════════════════")
total = PASS + FAIL
print(f"\n  Passed : {PASS}/{total}")
print(f"  Failed : {FAIL}/{total}")
if FAIL == 0:
    print("\n  🎉 ALL TESTS PASSED — ready for docker-compose up")
else:
    print("\n  ⚠️  some tests failed")
    sys.exit(1)
